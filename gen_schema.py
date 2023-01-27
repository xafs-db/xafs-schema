#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
   gen_schema.py
"""

import os
import argparse
import collections
import openpyxl
import json
import codecs
import datetime

from logging import getLogger, StreamHandler, DEBUG

logger = getLogger(__name__)

handler = StreamHandler()
handler.setLevel(DEBUG)
logger.setLevel(DEBUG)
logger.addHandler(handler)
logger.propagate = False


class G:
    schema_file = 'xafs-schema.json'
    schema_file_strict = 'xafs-schema-strict.json'


# date, datetimeの変換関数
def json_serial(obj):
    # 日付型の場合には、文字列に変換します
    if isinstance(obj, (datetime.datetime, datetime.date)):
        return obj.isoformat()
    # 上記以外はサポート対象外.
    raise TypeError ("Type %s not serializable" % type(obj))


def dump_json(out_dict, filename, debug=False):
    """
    utility function  (JSON出力)
    """

    if debug:
        if filename is not None:
            logger.info("--> filename = {}".format(filename))
        logger.info(json.dumps(out_dict, indent=4,
                               separators=(',', ':'), ensure_ascii=False))

    with codecs.open(filename, "w", "utf-8") as f:
        json.dump(out_dict, f, indent=4, separators=(
            ',', ':'), ensure_ascii=False,default=json_serial)


def deepupdate(dict_base, other):
    """
    utility function (辞書コピー)
    """
    for k, v in other.items():
        if isinstance(v, collections.abc.Mapping) and k in dict_base:
            deepupdate(dict_base[k], v)
        else:
            try:
                type2 = type(dict_base[k])
            except:
                type2 = None
            if type(v) == list and type2 == list:
                dict_base[k].extend(v)
            else:
                dict_base[k] = v


def convert_type(data_type):
    """
    utility function (データ型変換)
    """

    # elastic  --> json schema
    d = {}
    d["keyword"] = "string"
    d["text"] = "string"
    d["boolean"] = "boolean"  # not used
    d["integer"] = "integer"
    # any numeric type (integer, floating point numbers)
    d["double"] = "number"
    d["array"] = "array"
    d["date"] = "string"  # need to assign "date-time" in format field

    return d.get(data_type)


def get_parent_key_list(key, flag_flat_key=False):
    """
    utility function
    """

    key_list = [key]
    while 1:
        x = max(key.rfind("@"), key.rfind("."))
        if flag_flat_key:
            x = key.rfind(".")
        key_base = key[:x]
        if x < 0 or key_base == "":
            break
        key_list.append(key_base)
        key = key_base

    key_list.reverse()
    return key_list


def get_key_base_and_child(key, flag_flat_key=False):
    """
    utility function
    """

    x = max(key.rfind("@"), key.rfind("."))
    if flag_flat_key:
        x = key.rfind(".")

    key_base = "root"
    key_child = key[1:]
    if flag_flat_key:
        key_child = key
    if x > 0:
        key_base = key[:x]
        key_child = key[x+1:]

    return key_base, key_child


def get_items_from_sheet(sheet, flag_strict=False):
    """
    utility function (エクセルシートからアイテム取得)
    """

    col_key = 1
    col_data_type = 3
    col_desc_ja = 4
    col_desc = 5
    col_example = 6
    col_supplement = 7
    col_enum = 8
#    col_local     = 9

    required = []

    vdict = collections.OrderedDict()

    for i in range(2, sheet.max_row + 1):
        # ... check key
        key = sheet.cell(row=i, column=col_key).value
        if key is None:
            continue

        each_dict = collections.OrderedDict()
        each_dict["$id"] = "#key/{}".format(key)

        # ... check type
        data_type = sheet.cell(row=i, column=col_data_type).value
        data_type_converted = convert_type(data_type)
        if data_type_converted is not None:
      
            if flag_strict:

                if data_type != "date":
                    each_dict["type"] = data_type_converted
                else:
                    each_dict["anyOf"] = [{"type": "string",
                                          "pattern":  "^\d{4}\-(0?[1-9]|1[012])\-(0?[1-9]|[12][0-9]|3[01])( ([01][0-9]|2[0-3]):[0-5][0-9](:[0-5][0-9])?)(\.\d+)?$"},
                                          {"type": "string", "format": "date"},
                                          {"type": "string", "format": "date-time"}]

            else:
                if data_type_converted in ["array"]:
                    each_dict["type"] = data_type_converted
                elif data_type != "date":
                    each_dict["type"] = ["string","number","null"]
                else:
                    each_dict["anyOf"] = [{"type": "string",
                                          "pattern":  "^\d{4}\-(0?[1-9]|1[012])\-(0?[1-9]|[12][0-9]|3[01])( ([01][0-9]|2[0-3]):[0-5][0-9](:[0-5][0-9])?)(\.\d+)?$"},
                                          {"type": "string", "format": "date"},
                                          {"type": "string", "format": "date-time"},
                                          {"type": "null"}]


        # ... check description
        value_ja = sheet.cell(row=i, column=col_desc_ja).value
        value = sheet.cell(row=i, column=col_desc).value
        supplement_value = sheet.cell(row=i, column=col_supplement).value
        enum_value = sheet.cell(row=i, column=col_enum).value
        if value_ja or value:
            if value_ja is None:
                value_ja = ""
            if value is None:
                value = ""
            description = "{}\n{}".format(value_ja, value)
            if supplement_value is not None:
                description += "\nSupplement:\n{}".format(supplement_value)
            if enum_value is not None:
                description += "\nEnum:\n{}".format(enum_value)            

            each_dict["description"] = description

        # ... add title
        each_dict["title"] = key

        # ... add items for array
        if data_type == "array":
            each_dict["items"] = collections.OrderedDict()
            each_dict["items"]["type"] = "object"
            each_dict["items"]["properties"] = collections.OrderedDict()

        # ... check examples
        value = sheet.cell(row=i, column=col_example).value
        if value:
            each_dict["examples"] = [value]

        # ... check enum
        value = sheet.cell(row=i, column=col_enum).value
        if value:
            vsplit = value.split(",")
            enum_list = []
            for v in vsplit:
                enum_list.append(v.strip())

            if len(enum_list) > 0:
                each_dict["enum"] = enum_list

        # ... check local
#        value = sheet.cell(row=i, column=col_local).value
#        if value is None or value.lower() != "primary":
#            each_dict["tag"] = "local"
#        else:
#            each_dict["tag"] = "primary"

        if data_type == "array":
            each_dict["items"]["additionalProperties"] = False

        vdict[key] = each_dict

    return vdict


def get_required_from_sheet(sheet, flag_flat_key=False):
    """
    utility function
    """

    col_key = 1
    col_option = 2

    required = []

    vdict = collections.OrderedDict()

    for i in range(2, sheet.max_row + 1):
        # ... check key
        key = sheet.cell(row=i, column=col_key).value
        if key is None:
            continue

        key_base, key_child = get_key_base_and_child(key,
                                                     flag_flat_key=flag_flat_key)

        # ... check option
        value = sheet.cell(row=i, column=col_option).value
        if value == "required":
            if key_base not in vdict:
                vdict[key_base] = []

            vdict[key_base].append(key_child)

    return vdict


def get_items(items_dict, required_dict, flag_flat_key=False):
    """
    utility function
    """

    items_all_dict = collections.OrderedDict()

    if flag_flat_key is False:
        items_dict["@local@any"] = dict(type="object")

    key_all_list = list(items_dict.keys())
    key_list = list(items_dict.keys())

    for key in key_list:
        parent_key_list = get_parent_key_list(key,
                                              flag_flat_key=flag_flat_key)
        for parent_key in parent_key_list:

            key_base, key_child = get_key_base_and_child(parent_key,
                                                         flag_flat_key=flag_flat_key)

            data_type = items_dict.get(key_base, {}).get("type", "object")

            if key_base not in key_all_list:
                key_all_list.append(key_base)
                each_dict = collections.OrderedDict()
                if key_base == "root":
                    each_dict["$schema"] = "http://json-schema.org/draft/2020-12/schema"
                    each_dict["type"] = "object"
                    each_dict["title"] = "XAFS schema"
                    each_dict["description"] = "XAFS schema comprises the entire JSON document."
                elif data_type == "object":
                    each_dict["$id"] = "#/key/{}".format(key_base)
                    each_dict["type"] = "object"
                else:
                    raise Exception("illegal format for {}".format(key_base))

                each_dict["properties"] = collections.OrderedDict()
                if key_base != "@local":
                    each_dict["additionalProperties"] = False
                else:
                    each_dict["additionalProperties"] = True

                items_all_dict[key_base] = each_dict

            if data_type == "object":
                items_all_dict[key_base]["title"] = key_base
                items_all_dict[key_base]["properties"][key_child] = {
                    "$ref": "#/definitions/{}".format(parent_key)}
            elif data_type == "array":
                items_all_dict[key_base]["title"] = key_base
                items_all_dict[key_base]["items"]["properties"][key_child] = {
                    "$ref": "#/definitions/{}".format(parent_key)}
            else:
                raise Exception("illegal format for {}".format(key_base))

            if key_base in required_dict.keys() and \
               "required" not in items_all_dict[key_base]:
                items_all_dict[key_base]["required"] = required_dict[key_base]

                parent_key_base_list = get_parent_key_list(key_base,
                                                           flag_flat_key=flag_flat_key)
                for parent_key_base in parent_key_base_list:
                    if parent_key_base == "root":
                        continue

                    key_base2, key_child2 = get_key_base_and_child(parent_key_base,
                                                                   flag_flat_key=flag_flat_key)
                    if "required" not in items_all_dict[key_base2]:
                        items_all_dict[key_base2]["required"] = []
                    if key_child2 not in items_all_dict[key_base2]["required"]:
                        items_all_dict[key_base2]["required"].extend(
                            [key_child2])

        items_all_dict[key] = items_dict[key]

    # ... delete items for array with no components
    key_list_array_with_no_items = []
    for key, value in items_all_dict.items():
        v_type = value.get("type")
        if v_type != "array":
            continue
        n_properties = len(
            list(value.get("items", {}).get("properties", {}).keys()))
        if n_properties == 0:
            key_list_array_with_no_items.append(key)

    for key in key_list_array_with_no_items:
        del items_all_dict[key]["items"]

    return items_all_dict


def get_schema(filename_input, flag_required=True, flag_flat_key=False, flag_strict=False):
    """
    エクセル形式のメタデータスキーマより jsonschemaを作成

    Args:
        filename_input (str): _description_
        flag_required (bool, optional): required オプションを有効化フラグ. Defaults to True.
        flag_flat_key (bool, optional): BENTEN形式のメタデータ(flat構造)の有効化フラグ. Defaults to False.
        flag_strict (bool, optional): データ型チェックの有効化フラグ. Defaults to False.

    Returns:
        disct: jsonshemaの辞書
    """

    # ... get items_dict, required_dict from excel file
    wb = openpyxl.load_workbook(filename_input, data_only=True)

    items_dict = collections.OrderedDict()
    required_dict = collections.OrderedDict()
    for sheet_name in wb.sheetnames:
        if sheet_name in ['README']:
            continue

        deepupdate(items_dict, get_items_from_sheet(
            wb[sheet_name], flag_strict=flag_strict))

        if flag_required:
            deepupdate(required_dict, get_required_from_sheet(wb[sheet_name],
                                                              flag_flat_key=flag_flat_key))

    items_all_dict = get_items(items_dict, required_dict,
                               flag_flat_key=flag_flat_key)

    root_dict = items_all_dict['root']
    del items_all_dict['root']

    schema_dict = root_dict
    schema_dict['definitions'] = items_all_dict

    return schema_dict


if __name__ == "__main__":

    parser = argparse.ArgumentParser(
        description='generate xafs jsonschema from schema excel file')
    parser.add_argument('schema_excel_filename', help='schema excel filename')
    args = parser.parse_args()

    schema_excel_filename = args.schema_excel_filename

    logger.info(f'... input {schema_excel_filename}')

    output_dir = './schema'
    os.makedirs(output_dir, exist_ok=True)

#    flag_required = False
    flag_required = True

    schema_dict_strict = get_schema(schema_excel_filename,
                                    flag_required=flag_required, flag_strict=False)
    filename_schema = f'{output_dir}/{G.schema_file}'
    logger.info(f'... output {filename_schema}')

    dump_json(schema_dict_strict, filename_schema)

    schema_dict = get_schema(schema_excel_filename,
                             flag_required=flag_required, flag_strict=True)
    filename_schema = f'{output_dir}/{G.schema_file_strict}'

    logger.info(f'... output {filename_schema}')

    dump_json(schema_dict, filename_schema)
